from datetime import datetime
import uuid
from django.http import JsonResponse
from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.response import Response
# from adminapp.api.serializers import AdminSerializer
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from customer.api.sendmobileotp import send_otp
from customer.models import MobileOTP, User
from rest_framework import status
from rest_framework_simplejwt.serializers import RefreshToken
from superadmin.api.serializers import AdminSubscriptionSerializer, SuperadminLoginSerializer, TrackerListInfoSerializer, TrackerListSerializer, VendorDetailSerializer
import re
from adminapp.models import AdminTripDetails, Trip,AdminProfile
from rest_framework import generics
from driverapp.models import TrackerDeviceIntergrations
from driverapp.api.serializers import TrackerDeviceIntergrationSerializer
from superadmin.api.serializers import TrackerDeviceIntergrationDataSerializer
from superadmin.models import AdminPayment, AdminSubscription, TrackerDeviceInfo
from driverapp.api.serializers import TrackerDeviceIntergrationAllSerializer
from customer.models import Coupon
from customer.api.serializers import Coupon_Serializer, Coupon_List_Serializer


class LoginViewSet(viewsets.ModelViewSet, TokenObtainPairView):
    serializer_class = SuperadminLoginSerializer
    permission_classes = (AllowAny,)
    def create(self, request):
        password = request.data.get('password')
        email = request.data.get('email')
        
        if not email:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "email / passowrd is not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
           
            
        if email:
            mail_regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
            if not (re.search(mail_regex, email)):
                response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "Invalid email id",
                'result': {},
                'additional_data': {},
                }
                return Response(response_payload)
        if not password:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Password is not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
                
        try:
            user=User.objects.get(email=email,role='superadmin',is_active=True)
        except Exception as e:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please enter valid Email or Contact to Admin your a/c is disable',
                'result': {},
                'additional_data': {},
            }

            return Response(response_payload)
        if user.email !=email:
            response_payload = {
            'is_authenticated': False,
            'status':status.HTTP_400_BAD_REQUEST,
            'messages': 'Please enter valid Email',
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)
        if not user.check_password(password):
            response_payload = {
            'is_authenticated': False,
            'status':status.HTTP_400_BAD_REQUEST,
            'messages': 'Please enter valid Passoword',
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)
        
        refresh_token=RefreshToken.for_user(user)
        access_token = str(refresh_token.access_token)
        refresh_token = str(refresh_token)     
        # profile,created=AdminProfile.objects.get_or_create(user=user)
        login_serializer=SuperadminLoginSerializer(user)
        login_data=login_serializer.data
        
        response_payload = {
            'is_authenticated': True,
            'status':status.HTTP_200_OK,
            "messages":"Login Successful",
            'result': login_data,
            'access_token': access_token,
            'refresh_token': refresh_token,    
            'additional_data': {},
        }
        return Response(response_payload)
    
class SuperAdminRegisterView(viewsets.ViewSet):
    permission_classes = (AllowAny,)
    def create(self, request):
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        mobile = request.data.get('mobile')
        country_code = request.data.get('country_code')
        email = request.data.get('email')
        address=request.data.get('address')
        password = request.data.get('password')
        confirmPassword = request.data.get('confirm_password') or request.data.get('confirmPassword')
        role=request.data.get('role')
        if not first_name:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "First name not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not last_name:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Last name not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not mobile:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Mobile number not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not country_code:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Country Code not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not email:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Email ID not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        # if not role:
        #     response_payload = {
        #         'is_authenticated': False,
        #         'status':status.HTTP_400_BAD_REQUEST,
        #         'messages': "role not provided",
        #         'result': {},
        #         'additional_data': {},
        #     }
        #     return Response(response_payload)    
        if not re.match('([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+',email):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Email ID not valid",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not address:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "address not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
        if not password:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "Password not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not confirmPassword:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "confirm Password  not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)

        if password != confirmPassword:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  ' Password did not match ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        user_obj=User.objects.filter(email=email).exists()
        if user_obj:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': "this mail already exists",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        mobile_obj=User.objects.filter(mobile=mobile).exists()
        if mobile_obj:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': "Admin with this mobile already exists",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        user_object = User()
        user_object.first_name=first_name
        user_object.last_name=last_name
        user_object.email = email.lower()
        user_object.set_password(password)
        user_object.mobile = mobile
        user_object.country_code = country_code
        user_object.role="superadmin"
        user_object.is_staff=True
        user_object.is_active=True
        user_object.save()
        login_serializer=SuperadminLoginSerializer(user_object)
        login_data=login_serializer.data
        
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_201_CREATED,
            'messages': 'Registration successfully',
            'result': login_data,
            'additional_data': {},
        }
        return Response(response_payload)
class GetCustomerView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        customer_ids=User.objects.filter(role='customer').values_list('id',flat=True)
        payload=[]
        for x in customer_ids:
            queryset=User.objects.filter(id=x,role='customer')
            for y in queryset:
                payload.append({
                    'first_name':y.first_name,
                    'last_name':y.last_name,
                    'mobile':y.mobile,
                    'type':y.type,
                    'GST':y.gst,
                    'address':y.address,
                    'status':y.is_active
                })
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'customer details fetched succesfuly',
            'result': payload,
            'additional_data': {},  
        }
        return Response(response_payload)
class CustomerTripDetailsView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        customer_ids=User.objects.filter(role='customer').values_list('id')
        payload=[]
        for x in customer_ids:
            customer_id=x[0]
            queryset=Trip.objects.filter(user_id=x)
            for y in queryset:
                payload.append({
                    'customer_id':customer_id,
                    'starting_point':y.starting_point,
                    'destination':y.destination,
                    'material':y.material.material_type,
                    'truck_type':y.truck_type,
                    'date':y.starting_date,
                    'mobile':y.user.mobile,
                    'payment_status':y.payment_status,
                    'quotes':y.other_constraints,
                    'constraints':{"temprature":y.submaterial.temprature,"humidity":y.submaterial.humidity,"tilt":y.submaterial.tilt,"ambient_light":y.submaterial.ambient_light,"pitch_angle":y.submaterial.pitch_angle,"roll_angle":y.submaterial.roll_angle}                    
                })
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'customer details fetched succesfuly',
            'result': payload,
            'additional_data': {},  
        }
        return Response(response_payload)
    
class AdminListView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        admin_ids=User.objects.filter(role='admin').values_list('id',flat=True)
        payload=[]
        for x in admin_ids:
            queryset=User.objects.filter(id=x,role='admin')
            for y in queryset:
                try:
                    admin=AdminProfile.objects.get(user_id=x)
                except:
                    admin=None
                if admin:  
                    company_name = admin.company_name
                    address=admin.address
                    gst=admin.gst
                else:
                    pass
                try:
                    admin_device=TrackerDeviceInfo.objects.filter(admin_id=x).values('device')
                except:
                    admin_device=None
                # tracker_device=[]
                # for device in admin_device:
                #     tracker_device.append(device.device)
                # if tracker_device:
                #     device=tracker_device
                #     # print(tracker_device)
                # else:
                #     device=None
                payload.append({
                    'name':company_name,
                    'admin_id':y.id,
                    'mobile':y.mobile,
                    'type':y.type,
                    'GST':gst,
                    'device':admin_device,
                    'address':address,
                    'subscription':y.subscription,
                    'status':y.is_active
                    
                })
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'Admin details fetched succesfuly',
            'result': payload,
            'additional_data': {},  
        }
        return Response(response_payload)
    


class AdminUpdateAPI(viewsets.ViewSet):
    # permission_classes=(IsAuthenticated,)
    def update(self,request):
        admin_id = request.data.get('admin_id')
        address = request.data.get('address')
        village = request.data.get('village')
        city = request.data.get('city')
        state = request.data.get('state')
        pincode = request.data.get('pincode')
        company_name = request.data.get('company_name')
        gst = request.data.get('gst')
        pan_number = request.data.get('pan_number')

        try:
            admin = AdminProfile.objects.get(user_id=admin_id)
        except AdminProfile.DoesNotExist:
            return Response ({
            'is_authenticated': False,
            'status': status.HTTP_404_NOT_FOUND,
            'messages': 'Please share correct admin id or this admin not found',
            'result': "",
            'additional_data': {},  
        })

        if address is not None:
            admin.address = address

        if village is not None:
            admin.village = village

        if city is not None:
            admin.city = city
        
        if state is not None:
            admin.state = state
        
        if pincode is not None and not len(pincode)>6:
            admin.pincode = pincode
        else:
            return Response({
            'is_authenticated': False,
            'status': status.HTTP_404_NOT_FOUND,
            'messages': 'Pincode Cannot be more than 6 Digits',
            'result': "",
            'additional_data': {},  
        })

        
        if company_name is not None:
            admin.company_name = company_name

        if gst is not None:
            admin.gst = gst
        
        if pan_number is not None and not len(pan_number)>10:
            admin.pan_number = pan_number
        
        else:
            return Response({
            'is_authenticated': False,
            'status': status.HTTP_404_NOT_FOUND,
            'messages': 'Pan Number Cannot be More than 10 Characters',
            'result': "",
            'additional_data': {},  
        })

        admin.save()
        
        
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'Admin details Updated succesfuly',
            'result': "",
            'additional_data': {},  
        }
        return Response(response_payload)    
    

    
class TrackerDeviceGeneric(generics.GenericAPIView):
    serializer_class=TrackerDeviceIntergrationDataSerializer
    permission_classes = (IsAuthenticated,)
    def get(self, request):
        queryset = TrackerDeviceIntergrations.objects.filter(type='Device Information',status=True).distinct('device')[::-1]
        serializer_data = TrackerDeviceIntergrationSerializer(queryset,many=True)
        pay_load=serializer_data.data
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': "",
            'result': pay_load,
            'additional_data': {},
        }
        return Response(response_payload)
    def post(self,request):
        # type=request.data.get('type')
        firmware_version=request.data.get('firmware_version')
        hardware_version=request.data.get('hardware_version')
        # serial_no=request.data.get('serial_no')
        device=request.data.get('device')
        
        # if not type:
        #     response_payload = {
        #     'is_authenticated': False,
        #     'status': status.HTTP_400_BAD_REQUEST,
        #     'messages': "type is not provided",
        #     'result': {},
        #     'additional_data': {},
        #     }
        #     return Response(response_payload)
        if not firmware_version:
            response_payload = {
            'is_authenticated': False,
            'status': status.HTTP_400_BAD_REQUEST,
            'messages': "firmware version not provided ",
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)
        if not hardware_version:
            response_payload = {
            'is_authenticated': False,
            'status': status.HTTP_400_BAD_REQUEST,
            'messages': "type is not hardware version ",
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)
        if not device:
            response_payload = {
            'is_authenticated': False,
            'status': status.HTTP_400_BAD_REQUEST,
            'messages': "device is not provided ",
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)  
        # if not serial_no:
        #     response_payload = {
        #     'is_authenticated': False,
        #     'status': status.HTTP_400_BAD_REQUEST,
        #     'messages': "serial no is not provided ",
        #     'result': {},
        #     'additional_data': {},
        #     }
        #     return Response(response_payload)
    
        if TrackerDeviceInfo.objects.filter(device=device).exists():
            response_payload = {
            'is_authenticated': False,
            'status': status.HTTP_400_BAD_REQUEST,
            'messages': "Device Id already exist",
            'result': {},
            'additional_data': {},
            }
            return Response(response_payload)
        else:
            device_obj=TrackerDeviceInfo()
            device_obj.type="Device Information"
            device_obj.firmware_version=firmware_version
            device_obj.hardware_version=hardware_version
            device_obj.serial_no="0001"
            device_obj.device=device
            device_obj.save()
            tracker_obj=TrackerDeviceIntergrations()
            tracker_obj.type="Device Information"
            tracker_obj.firmware_version=firmware_version
            tracker_obj.hardware_version=hardware_version
            tracker_obj.serial_no="0001"
            tracker_obj.device=device
            tracker_obj.save()
            tracker_serialiser= TrackerDeviceIntergrationDataSerializer(tracker_obj)
            pay_load=tracker_serialiser.data
            response_payload = {
                'is_authenticated': True,
                'status': status.HTTP_200_OK,
                'messages': "Tracker added succesfully",
                'result': pay_load,
                'additional_data': {},
                }
            return Response(response_payload)
       
       
class TrackerHistoryView(generics.GenericAPIView):
    queryset = TrackerDeviceIntergrations.objects.all()
    serializer_class = TrackerDeviceIntergrationAllSerializer
    # pagination_class = PagePagination
   
    def get(self, request,slug=None):
        # slug=request.GET.get('slug')
        queryset = TrackerDeviceIntergrations.objects.filter(device=slug).order_by('-id')
        serializer = TrackerDeviceIntergrationAllSerializer(queryset, many=True)
        pay_load=serializer.data
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': [],
            'result': pay_load,
            'additional_data': {},
        }
        return Response(response_payload)
    
class TrackerDeviceDetails(generics.GenericAPIView):
    queryset = TrackerDeviceIntergrations.objects.all()
    serializer_class = TrackerDeviceIntergrationAllSerializer
    # pagination_class = PagePagination
    permission_classes=(IsAuthenticated,)
    def get(self, request, slug=None):
        try:
            queryset = TrackerDeviceIntergrations.objects.filter(device=slug).order_by('-id')[0]
        except:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Device ID does not exist',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
        serializer = TrackerDeviceIntergrationAllSerializer(queryset)
        pay_load=serializer.data
        response_payload = {
        'is_authenticated': True,
        'status': status.HTTP_200_OK,
        'messages': [],
        'result': pay_load,
        'additional_data': {},
        }
        return Response(response_payload)
       
class TrackerListInfoView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        tracker=TrackerDeviceInfo.objects.all().distinct('device')
        serializer=TrackerListInfoSerializer(tracker,many=True)
        # tracker=TrackerDeviceIntergrations.objects.values('device','admin','tracker_availability').distinct().order_by('device')
        response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'tracker list fetched succesfully',
            'result': serializer.data,
            'additional_data': {},  
        }
        return Response(response_payload) 
       
class TrackerAssignView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def create(self,request):
        admin_id=request.data.get('admin_id')
        device=request.data.get('device')
        if not admin_id:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide the admin id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not device:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide the device id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        try:
            uuid_obj = uuid.UUID(admin_id) 
        except ValueError:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide a valid admin id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
        if TrackerDeviceInfo.objects.filter(device=device).exists() and User.objects.filter(id=admin_id,role='admin').exists():
            TrackerDeviceInfo.objects.filter(device=device).update(admin_id=admin_id,tracker_availability='False')
            # TrackerDeviceIntergrations.objects.filter(device=tracker_id,type='Device Information').update(admin_id=admin_id)
            response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_200_OK,
            'messages': 'Tracker assigned to admin succesfully',
            'result': {},
            'additional_data': {},  
            }
            return Response(response_payload)  
        else:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'This tracker id or admin id does not exists',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
# This is for dropdown while assigning tracker to admin             
class TrackerListAssignView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        tracker_obj=TrackerDeviceInfo.objects.filter(tracker_availability='True')
        if tracker_obj:
            serializer=TrackerListSerializer(tracker_obj,many=True)
            response_payload = {
                'is_authenticated': True,
                'status': status.HTTP_200_OK,
                'messages': 'Following trackers are available to assign',
                'result': serializer.data,
                'additional_data': {},  
                }
            return Response(response_payload)         
        if not tracker_obj:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'No trackers available currently',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)       
                   

class AdminStatusUpdate(generics.GenericAPIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        admin_id = request.data.get('admin_id')

        if not admin_id or not re.fullmatch("^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",admin_id):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_401_UNAUTHORIZED,
                'messages': 'Valid Admin ID not found',
            }
            return Response(response_payload)     
        
        try:
            user = User.objects.get(id=admin_id, role='admin')
        except User.DoesNotExist:
            user = None

        if user is not None: 
            if not user.is_active==True:
                user.is_active = True
                user.save()
                response_payload = {
                    'is_authenticated': True,
                    'status': status.HTTP_200_OK,
                    'messages': "Admin Status Set to True",
                    'result': {},
                    'additional_data': {},
                }
                return Response(response_payload)
            else:
                response_payload = {
                        'is_authenticated': False,
                        'status': status.HTTP_200_OK,
                        'messages': "Admin is already Verified",
                        'result': {},
                        'additional_data': {},
                    }
                return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_401_UNAUTHORIZED,
                'messages': 'Admin Does not exist',
            }
            return Response(response_payload)
        

class DeactivateAdmin(generics.GenericAPIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        admin_id = request.data.get('admin_id')

        if not admin_id or not re.fullmatch("^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",admin_id):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_401_UNAUTHORIZED,
                'messages': 'Valid Admin ID not found',
            }
            return Response(response_payload)
        
        try:
            user = User.objects.get(id=admin_id, role='admin')
        except User.DoesNotExist:
            user = None

        if user is not None: 
            if user.is_active==True:
                user.is_active = False
                user.save()
                response_payload = {
                    'is_authenticated': True,
                    'status': status.HTTP_200_OK,
                    'messages': "Admin Deactivated",
                    'result': {},
                    'additional_data': {},
                }
                return Response(response_payload)
            else:
                response_payload = {
                        'is_authenticated': False,
                        'status': status.HTTP_200_OK,
                        'messages': "Admin is already Deactive",
                        'result': {},
                        'additional_data': {},
                    }
                return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_401_UNAUTHORIZED,
                'messages': 'Admin Does not exist',
            }
            return Response(response_payload)
      
        

class SendOtpView(viewsets.ViewSet):
    def create(self,request):
        # import pdb; pdb.set_trace()

        mobile = request.data.get('mobile')
        country_code = request.data.get('country_code')
        if not mobile:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'mobile number not provided',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not country_code:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'country code not provided',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        
        try:
            user_object = User.objects.get(mobile=mobile,country_code=country_code,role='superadmin')
        except:
            
            response_payload = {
            'is_authenticated': False,
            'status': status.HTTP_400_BAD_REQUEST,
            'messages': 'Mobile number or country code not valid',
            'result': [],
            'additional_data': {},
            }
            return Response(response_payload)
        if user_object:
            otp_object,created =MobileOTP.objects.get_or_create(user=user_object)
            otp_object.mobile=mobile
            otp_object.otp=send_otp(mobile,country_code)["otp"]
            otp_object.time=int(float((datetime.now().timestamp())))
            otp_object.save()
            response_payload = {
            'is_authenticated': True,
            'status': status.HTTP_201_CREATED,
            "mobile_no":str(user_object.mobile),
            'messages': 'OTP send succesfully on your mobile number',
            'result': [],
            'additional_data': {},
            }
            return Response(response_payload)
        
class VerifyOtpView(viewsets.ViewSet):
    def create(self,request):
        # import pdb; pdb.set_trace()
        otp = request.data.get('otp')
        mobile = request.data.get('mobile')
        if not mobile:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'mobile number not provided',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not otp:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'otp not provided',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        try:
            mobile_otp = MobileOTP.objects.get(user__mobile=mobile)
        except:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Mobile number not valid',
                'result': [],
                'additional_data': {},
            }

            return Response(response_payload)
        if  mobile_otp.otp==otp:
            current_time=datetime.now().timestamp()
            if int(float(current_time))-int(mobile_otp.time)<300:
                mobile_otp.otp_verify = True
                mobile_otp.count +=1
                mobile_otp.save()
                response_payload = {
                    'is_authenticated': True,
                    'status':status.HTTP_200_OK,
                    "messages":"Succesfully verified your OTP",
                    'result': [],
                    'additional_data': {},
                }
                return Response(response_payload)
            else:
                response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'your otp number has been Expire  please login again',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please enter valid OTP',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
class ResetPasswordView(viewsets.ViewSet):
    def create(self,request):
        mobile=request.data.get('mobile')
        new_password=request.data.get('new_password')
        confirm_password=request.data.get('confirm_password')
        if not mobile:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'please provide mobile number',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not new_password:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'please enter your new password',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(mobile=mobile,role='superadmin'):
            if new_password==confirm_password:
                user=User.objects.get(mobile=mobile)
                user.set_password(new_password)
                user.save()
                response_payload = {
                'is_authenticated': True,
                'status':status.HTTP_200_OK,
                'messages': 'password updated succesfully',
                'result': [],
                'additional_data': {},
            }
                return Response(response_payload)
            else:
                response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'both passwords dont match',
                'result': [],
                'additional_data': {},
            }
                return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'mobile number not registered',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        
class AssignAdminTrip(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def create(self,request):
        trip_id=request.data.get('trip_id')
        admin_id=request.data.get('admin_id')
        if not trip_id:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide the trip id',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not admin_id:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide the admin id',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not re.fullmatch("[a-f0-9]{8}-?[a-f0-9]{4}-?4[a-f0-9]{3}-?[89ab][a-f0-9]{3}-?[a-f0-9]{12}",trip_id):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Enter valid trip id',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if not re.fullmatch("[a-f0-9]{8}-?[a-f0-9]{4}-?4[a-f0-9]{3}-?[89ab][a-f0-9]{3}-?[a-f0-9]{12}",admin_id):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Enter valid admin id',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        if Trip.objects.filter(id=trip_id,trip_status='requested').exists():
            if User.objects.filter(id=admin_id,role='admin',is_active=True):
                AdminTripDetails.objects.create(admin_id=admin_id,trip_id=trip_id,trip_status='requested')
                response_payload = {
                    'is_authenticated': True,
                    'status':status.HTTP_200_OK,
                    'messages': 'Trip assigned to admin succesfully',
                    'result': [],
                    'additional_data': {},
                }
                return Response(response_payload)
            else:
                response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Admin id does not exists or is not active',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Trip id does not exists',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        

class CreateCouponCodeAPIView(generics.GenericAPIView):
    # permission_classes = (IsAuthenticated,)
    def post(self, request):
        percent_discount = request.data.get('percent_discount')
        coupon_amt =  request.data.get('coupon_amt')
        coupon_code =  request.data.get('coupon_code')


        if not coupon_code or len(coupon_code) < 6:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please enter valid coupon code or atleast 6 characters',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        
        if not percent_discount:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please enter percent of discount you want to offer',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        
        if not coupon_amt:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': 'Please enter coupon amount',
                'result': [],
                'additional_data': {},
            }
            return Response(response_payload)
        
        if not Coupon.objects.filter(coupon_code=coupon_code).exists():
            create_coupon = Coupon.objects.create(coupon_code=coupon_code, coupon_amt=coupon_amt, percent_discount=percent_discount, type='common')
            create_coupon.save()
            response_payload = {
                    'is_authenticated': True,
                    'status': status.HTTP_200_OK,
                    'messages': "Coupon Created Succesfully",
                    'result': 
                        {
                            "Coupon Code": coupon_code,
                            "Coupon Amount": coupon_amt
                        }
                    ,
                    'additional_data': {},
                }
            return Response(response_payload)
        else:
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "This Coupon Code Already Exist",
                    'result': {},
                    'additional_data': {},
                }
            return Response(response_payload)
        

class AssignCoupon(generics.GenericAPIView):
    # permission_classes = (IsAuthenticated,)

    def post(self, request):
        coupon_code = request.data.get('coupon_code')
        user_id = request.data.get('user_id')
        try:
            coupon = Coupon.objects.get(coupon_code=coupon_code)
        except Coupon.DoesNotExist:
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': f"{coupon_code} Coupon does not exist, Please create new Coupon",
                    'result': "",
                    'additional_data': {},
                }
            return Response(response_payload)
        
        try:
            get_user = User.objects.get(id=user_id, role='customer')
        except User.DoesNotExist: 
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "User Does Not Exist",
                    'result': "",
                    'additional_data': {},
                }
            return Response(response_payload)
        
        if not get_user.coupon_code==coupon_code:
            get_user.coupon_code=coupon_code
            get_user.save()
            coupon.user_id = user_id
            coupon.type = 'specific'
            coupon.save()
            response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': f"{coupon_code} Coupon Assigned to Given User",
                        'result': '',
                        'additional_data': {},
                    }
            return Response(response_payload)


class ListAllAvailableCouponsToApply(generics.GenericAPIView):
    # permission_classes = (IsAuthenticated,)
    def get(self,request):
        coupons = Coupon.objects.filter(user_id=None, type="common")
        serializer = Coupon_List_Serializer(coupons, many=True)
        response_payload = {
                    'is_authenticated': True,
                    'status': status.HTTP_200_OK,
                    'messages': "Available Common Coupons",
                    'result': serializer.data,
                    'additional_data': {},
                }
        return Response(response_payload)
        

class CustomerByNoOfOrders(generics.GenericAPIView):
    # permission_classes = (IsAuthenticated,)
    def get(self, request):
        trip_order = Trip.objects.filter(payment_status="paid")
        dict_order = {}
        id = []

        for i in trip_order:
                id.append(i.user_id)
        
        for i in id:
            if str(i) in dict_order:
                dict_order[str(i)] += 1    
            else:
                dict_order[str(i)] = 1 

        payload = []
        for j in dict_order:
            payload.append({
                "User id": j,
                "Order Count": dict_order[j]
              })

        response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': "Users with Number of Succesful Orders ",
                        'result': payload,
                        'additional_data': {},
                    }
        return Response(response_payload)
            
class AdminSubscriptionDetails(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        admin_id=request.data.get('admin_id')
        try:
            uuid_obj = uuid.UUID(admin_id) 
        except ValueError:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide a valid admin id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(id=admin_id,role='admin').exists():
            queryset=AdminSubscription.objects.filter(admin_id=admin_id)
            serializer=AdminSubscriptionSerializer(queryset,many=True)
            response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': "Following are the subscription details",
                        'result': serializer.data,
                        'additional_data': {},
                    }
            return Response(response_payload)   
        elif not admin_id:
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "Please enter the admin id",
                    'result': "",
                    'additional_data': {},
                }
            return Response(response_payload)
        else:
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "the admin id you entered is of wrong user",
                    'result': "",
                    'additional_data': {},
                }
            return Response(response_payload)

import csv
from openpyxl import load_workbook

class BulkAddTracker(generics.GenericAPIView):
    def post(self, request):
        if request.method == 'POST':            
            csv_file = request.FILES['tracker_info']
            # For .csv format
            if csv_file.name.endswith('.csv'):
                data = csv_file.read().decode('utf-8')
                data = data.splitlines()
                csv_data = csv.reader(data)
                csv_data = iter(csv_data)
                header = next(csv_data, None)
                if header:
                    for row in csv_data :
                        TrackerDeviceInfo.objects.get_or_create(
                            firmware_version=row[0],
                            hardware_version=row[1],
                            device = row[2],
                            serial_no = row[3],
                            type = "Device Information"
                        )
                    response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': "Tracker Data Added Successfully",
                        'result': "",
                        'additional_data': {},
                    }                        
                    return Response(response_payload)
                else:
                    response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "Please provide data in correct format",
                    'result': "",
                    'additional_data': {},
                }
                    return Response(response_payload)  

            # For .xlsx format
            else:           
                excel_file = request.FILES['tracker_info']          
                wb = load_workbook(excel_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    print(row)
                    firmware_version, hardware_version, device, serial_no = row
                    TrackerDeviceInfo.objects.get_or_create(firmware_version=firmware_version, hardware_version=hardware_version, device=device, type="Device Information", serial_no=serial_no)
                response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': "Tracker Data Added Successfully",
                        'result': "",
                        'additional_data': {},
                    }  
                return Response(response_payload)
        response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "Please provide data in correct format",
                    'result': "",
                    'additional_data': {},
                }
        return Response(response_payload)
class AdminPaymentDetailsView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)    
    def list(self,request):
        admin_id=request.GET.get('admin_id')
        payload=[]
        try:
            uuid_obj = uuid.UUID(admin_id) 
        except ValueError:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide a valid admin id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(id=admin_id,role='admin').exists():
            # admin_payments = AdminPayment.objects.filter(admin_id=admin_id)

        # Move trip query outside loop
            
            # admin_trips = Trip.objects.filter(admin_id=admin_id) 

            # for x in admin_payments:
            admin_trip_id=AdminPayment.objects.filter(admin_id=admin_id)
            # print('admin_trip_id',admin_trip_id)
            # trip_obj=Trip.objects.filter(id__in=admin_trip_id)

            # print('trip_id',trip_obj)
            for x in admin_trip_id:  
                payload.append({
                'amount_paid':x.amount,
                'status':x.payment_status,
                'payment_id':x.payment_id,
                'trip_id_show':[x.trip.trip_id_show if x.trip else ""][0],
                'trip_uuid':[x.trip.id if x.trip else ""][0],
                'to':[x.trip.destination if x.trip else ""][0],
                'from':[x.trip.starting_point if x.trip else ""][0]
                })
            response_payload = {
                        'is_authenticated': True,
                        'status': status.HTTP_200_OK,
                        'messages': "Following are the payment details",
                        'result': payload,
                        'additional_data': {},
                    }  
            return Response(response_payload)
                    
        # elif not admin_id:
        #     response_payload = {
        #             'is_authenticated': False,
        #             'status': status.HTTP_400_BAD_REQUEST,
        #             'messages': "Please enter the admin id",
        #             'result': "",
        #             'additional_data': {},
        #         }
        #     return Response(response_payload)
        else:
            response_payload = {
                    'is_authenticated': False,
                    'status': status.HTTP_400_BAD_REQUEST,
                    'messages': "the admin id you entered is of wrong user",
                    'result': "",
                    'additional_data': {},
                }
            return Response(response_payload)        
    
class SuperadminCreateAdminViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated)
    def create(self, request):
        company_name = request.data.get('company_name')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        mobile = request.data.get('mobile')
        country_code = request.data.get('country_code')
        email = request.data.get('email')
        gst = request.data.get('gst')
        pan_number = request.data.get('pan_number')
        address=request.data.get('address')
        password = request.data.get('password')
        confirmPassword = request.data.get('confirm_password') or request.data.get('confirmPassword')
        role=request.data.get('role')
        fcm_token=request.data.get('fcm_token')
        print("fcm token",fcm_token)
        if not company_name:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Company name is not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
        if AdminProfile.objects.filter(company_name=company_name).exists():
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "This company name is already registered with us",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not first_name:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "First name not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not last_name:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Last name not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not mobile:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Mobile number not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(mobile=mobile).exists():
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "User with this mobile number already exists",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not country_code:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Country Code not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not email:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Email ID not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(email=email).exists():
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "User with this email already exists",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
            
        if not re.match('([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+',email):
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Email ID not valid",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
            
        if not gst:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "gst not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not pan_number:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "pan number not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not len(pan_number)<=10:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages': "Pan number not is not greater than 10 digit",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not address:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "address not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        
        if not password:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "Password not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if not confirmPassword:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "confirm Password  not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)

        if password != confirmPassword:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  'Password did not match ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        user_object = User()
        user_object.first_name=first_name
        user_object.last_name=last_name
        user_object.email = email.lower()
        user_object.set_password(password)
        user_object.mobile = mobile
        user_object.country_code = country_code
        user_object.role="admin"
        user_object.is_active=True
        user_object.fcm_token=fcm_token
        user_object.save()
            
        profile, created=AdminProfile.objects.get_or_create(user=user_object, gst=gst,pan_number=pan_number,company_name=company_name)
        profile.save()
        response_payload = {
            'is_authenticated': True,
            'status':status.HTTP_200_OK,
            'messages': 'Admin Registration successfull',
            'result': {},
            'additional_data': {},
        }
        return Response(response_payload)
class VendorDetailsView(viewsets.ViewSet):
    permission_classes=(IsAuthenticated,)
    def list(self,request):
        admin_id=request.GET.get('admin_id')
        if not admin_id:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "Admin id not provided",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        try:
            uuid_obj = uuid.UUID(admin_id) 
        except ValueError:
            response_payload = {
                'is_authenticated': False,
                'status': status.HTTP_400_BAD_REQUEST,
                'messages': 'Please provide a valid admin id ',
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)
        if User.objects.filter(id=admin_id,role='admin').exists():
            queryset=AdminProfile.objects.get(user_id=admin_id)
            serializer=VendorDetailSerializer(queryset)
            response_payload = {
                'is_authenticated': True,
                'status':status.HTTP_200_OK,
                'messages': 'following are the admin details',
                'result':serializer.data,
                'additional_data': {},
            }
            return Response(response_payload)
        else:
            response_payload = {
                'is_authenticated': False,
                'status':status.HTTP_400_BAD_REQUEST,
                'messages':  "This Admin id is not registered with us",
                'result': {},
                'additional_data': {},
            }
            return Response(response_payload)